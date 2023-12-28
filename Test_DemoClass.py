from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
import pytest
import openpyxl
from constants import globalConstants as c
import json

class Test_DemoClass:
    #prefix => test_ 
    def setup_method(self): #her test başlangıcında çalışacak fonk
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 

    def teardown_method(self): # her testinin bitiminde çalışacak fonk
        self.driver.quit()

    def getData():
        excel = openpyxl.load_workbook(c.invalid_login_xlsx)
        sheet = excel["Sheet1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))

        return data
    
    def readInvalidDataFromJson():
        file = open("data/invalidLogin.json") 
        data = json.load(file)
        parameter = []

        for user in data['users']:
            username = user["username"]
            password = user["password"]
            parameter.append((username,password))

        return parameter

    @pytest.mark.parametrize("username,password",readInvalidDataFromJson())
    def test_invalid_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == c.USERNAME_PASSWORD_DONT_MATCH

    @pytest.mark.skip
    def test_valid_login(self):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        self.driver.execute_script("window.scrollTo(0,500)") #penceremin scrolunu aşağıya indiriyor,verdiğim koordinatlarla birlikte
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='add-to-cart-test.allthethings()-t-shirt-(red)']")))
        addToCart.click()
        remove = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='remove-test.allthethings()-t-shirt-(red)']")))
        assert remove.text == "Remove"

